from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import db
from app.models import FinanceRecord


def register(client, username: str = "tester", password: str = "Secret123!"):
    return client.post(
        "/auth/register",
        data={
            "username": username,
            "password": password,
            "confirm_password": password,
        },
        follow_redirects=True,
    )


def login(client, username: str = "tester", password: str = "Secret123!"):
    return client.post(
        "/auth/login",
        data={"username": username, "password": password},
        follow_redirects=True,
    )


def add_record(client, **overrides):
    payload = {
        "record_date": date(2024, 2, 1).isoformat(),
        "category": "เงินเดือน",
        "description": "เงินเดือนประจำ",
        "record_type": "income",
        "amount": "1200.50",
    }
    payload.update(overrides)
    return client.post("/", data=payload, follow_redirects=True)


def test_login_required_redirects(client):
    response = client.get("/", follow_redirects=False)
    assert response.status_code == 302
    assert "/auth/login" in response.headers["Location"]


def test_finance_flow(client, app):
    register_response = register(client)
    assert "สมัครสมาชิกสำเร็จ" in register_response.get_data(as_text=True)

    login_response = login(client)
    assert "เข้าสู่ระบบสำเร็จ" in login_response.get_data(as_text=True)

    dashboard_response = client.get("/")
    assert dashboard_response.status_code == 200

    add_response = add_record(client)
    assert "บันทึกข้อมูลเรียบร้อย" in add_response.get_data(as_text=True)

    with app.app_context():
        records = db.session.execute(db.select(FinanceRecord)).scalars().all()
    assert len(records) == 1
    record = records[0]
    assert record.category == "เงินเดือน"
    assert record.amount == pytest.approx(1200.50)

    dashboard_with_data = client.get("/")
    page_text = dashboard_with_data.get_data(as_text=True)
    assert "1,200.50" in page_text
    assert "คงเหลือ" in page_text

    download = client.get("/download")
    assert download.status_code == 200
    assert download.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(download.data))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][:5] == ("วันที่", "ประเภท", "หมวดหมู่", "รายละเอียด", "จำนวนเงิน")
    assert rows[1][0] == "2024-02-01"
    assert rows[1][1] == "รายรับ"
    assert rows[1][2] == "เงินเดือน"
    assert rows[1][4] == pytest.approx(1200.50)

    assert rows[-3][3] == "รวมรายรับ"
    assert rows[-3][4] == pytest.approx(1200.50)
    assert rows[-2][3] == "รวมรายจ่าย"
    assert rows[-2][4] == pytest.approx(0)
    assert rows[-1][3] == "คงเหลือ"
    assert rows[-1][4] == pytest.approx(1200.50)


def test_negative_amount_is_rejected(client, app):
    register(client)
    login(client)

    response = add_record(client, amount="-5")
    assert "จำนวนเงินต้องมากกว่า 0" in response.get_data(as_text=True)

    with app.app_context():
        assert db.session.execute(db.select(FinanceRecord.id)).first() is None
